import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
# pyinstaller --onefile job_manager.py


class JobApplicationManager:
    FILENAME = "JobApplications.xlsx"
    COLUMNS = ["POSITION", "COMPANY", "CITY", "LINK", "OUTCOME"]

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    white_side = Side(style='thin', color="FFFFFF")
    black_side = Side(style='thin', color="000000")
    header_border = Border(left=white_side, right=white_side, top=black_side, bottom=black_side)
    data_border = Border(left=black_side, right=black_side)

    outcome_styles = {
        "Waiting": {"font": Font(color="0000FF"), "fill": PatternFill(start_color="DCE6F1", fill_type="solid")},
        "Rejected": {"font": Font(color="FF0000"), "fill": PatternFill(start_color="FADBD8", fill_type="solid")},
        "Accepted": {"font": Font(color="00AA00"), "fill": PatternFill(start_color="D5F5E3", fill_type="solid")},
        "In Progress": {"font": Font(color="B8860B"), "fill": PatternFill(start_color="FFFACD", fill_type="solid")},
    }

    def __init__(self):
        self.setup_file()

    def setup_file(self):
        if not os.path.exists(self.FILENAME):
            wb = Workbook()
            ws = wb.active
            ws.append(self.COLUMNS)
            wb.save(self.FILENAME)

    def load_ws(self):
        wb = load_workbook(self.FILENAME)
        return wb, wb.active

    def save_wb(self, wb):
        wb.save(self.FILENAME)

    @staticmethod
    def sort_by_company(ws):
        rows = list(ws.iter_rows(min_row=2, max_col=5, values_only=True))
        rows.sort(key=lambda row: str(row[1]).lower() if row[1] else "")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.value = None

        for i, row_data in enumerate(rows, start=2):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

    def apply_formatting(self, ws):
        last_row = ws.max_row

        # Header formatting
        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.header_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data formatting
        for r in range(2, last_row + 1):
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.border = self.data_border
                if c == 4:  # Link column
                    cell.font = Font(bold=True)
                elif c == 5:  # Outcome column
                    val = str(cell.value)
                    if val in self.outcome_styles:
                        style = self.outcome_styles[val]
                        cell.font = style["font"]
                        cell.fill = style["fill"]

        # Adjust column widths (except link column)
        for col_idx in range(1, 6):
            if col_idx == 4:  # Skip "LINK"
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
                continue

            max_length = len(self.COLUMNS[col_idx - 1])
            for r in range(2, last_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if isinstance(val, str):
                    max_length = max(max_length, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    def add_job(self):
        position = input("\nEnter position title: ")
        company = input("Enter company name: ")
        city = input("Enter city: ")
        link = input("Enter application link: ")
        formatted_link = f'=HYPERLINK("{link}", "Here")'

        wb, ws = self.load_ws()
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=position)
        ws.cell(row=row, column=2, value=company)
        ws.cell(row=row, column=3, value=city)
        ws.cell(row=row, column=4, value=formatted_link)
        ws.cell(row=row, column=5, value="Waiting")

        self.sort_by_company(ws)
        self.apply_formatting(ws)
        self.save_wb(wb)
        print("✅ Job application added.\n")

    def list_jobs(self):
        wb, ws = self.load_ws()
        jobs = [(row[0], row[1], row[2], row[4]) for row in ws.iter_rows(min_row=2, max_col=5, values_only=True)]
        return jobs

    def edit_outcome(self):
        jobs = self.list_jobs()
        if not jobs:
            print("⚠️ No job applications found.\n")
            return

        print("\n📋 Job Applications:")
        for i, (position, company, city, outcome) in enumerate(jobs, start=1):
            print(f"[{i}] {company} - {position} ({city}) - [{outcome}]")

        try:
            choice = int(input("Select job number to update: "))
            if not (1 <= choice <= len(jobs)):
                print("❌ Invalid selection.\n")
                return
        except ValueError:
            print("❌ Please enter a valid number.\n")
            return

        print("\nNew outcome:")
        print("[1] Waiting")
        print("[2] Rejected")
        print("[3] Accepted")
        print("[4] In Progress")
        opt = input("Choice: ")

        mapping = {"1": "Waiting", "2": "Rejected", "3": "Accepted", "4": "In Progress"}
        outcome = mapping.get(opt)
        if not outcome:
            print("❌ Invalid outcome.\n")
            return

        wb, ws = self.load_ws()
        ws.cell(row=choice + 1, column=5, value=outcome)

        self.sort_by_company(ws)
        self.apply_formatting(ws)
        self.save_wb(wb)
        print("✅ Outcome updated.\n")

    def view_applications(self):
        jobs = self.list_jobs()
        if not jobs:
            print("\n📋 No job applications found.\n")
            return

        print("\n📋 Job Applications:")
        for i, (position, company, city, outcome) in enumerate(jobs, start=1):
            print(f"[{i}] {company} - {position} ({city}) - {outcome}")
        input("Press Enter to return to the main menu...\n")

    def main_menu(self):
        while True:
            print("=== Job Application Manager ===")
            print("[1] Add Job Application")
            print("[2] Edit Application Outcome")
            print("[3] View Applications")
            print("[4] Exit")

            choice = input("Select an option: ")
            if choice == "1":
                self.add_job()
            elif choice == "2":
                self.edit_outcome()
            elif choice == "3":
                self.view_applications()
            elif choice == "4":
                print("👋 Exiting. Goodbye!")
                break
            else:
                print("❌ Invalid choice.\n")


if __name__ == "__main__":
    manager = JobApplicationManager()
    manager.main_menu()
